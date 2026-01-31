import json
import os
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Protection, Border, Side


def _ensure_outputs_dir(base_dir: str) -> str:
    outputs_dir = os.path.join(base_dir, "outputs")
    os.makedirs(outputs_dir, exist_ok=True)
    return outputs_dir


def _load_relation_json(path: str) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def create_reverse_template(base_dir: str, student_count: int, relation_json_path: str = None) -> str:
    """
    创建逆向成绩模板。
    【修改】逆向模式必须依赖关系表，如果未提供则抛出异常。
    """
    # ===== 新增：强制检查关系表 =====
    if not relation_json_path or not os.path.exists(relation_json_path):
        raise ValueError("逆向模式必须先填写'课程考核与课程目标对应关系表'，请先完成关系表设置后再下载逆向模板。")
    
    data = _load_relation_json(relation_json_path)
    links = data.get("links", [])
    if not links:
        raise ValueError("关系表中未找到考核环节信息，请检查'课程考核与课程目标对应关系表'是否正确填写。")
    # ===== 强制检查结束 =====
    
    wb = Workbook()
    ws = wb.active
    ws.title = "逆向成绩模板"

    # 从关系表读取考核环节名称作为表头
    headers = ["姓名"] + [link.get("name", "").strip() for link in links]
    
    ws.append(headers)
    _apply_header_style(ws, 1, len(headers))
    _append_blank_rows(ws, student_count, len(headers))
    _protect_sheet(ws, editable_cols=list(range(1, len(headers) + 1)), header_rows=1)

    _apply_grid_borders(ws, ws.max_row, ws.max_column)

    output_path = os.path.join(_ensure_outputs_dir(base_dir), "逆向成绩模板.xlsx")
    wb.save(output_path)
    return output_path


def create_forward_template(base_dir: str, student_count: int, relation_json_path: str) -> str:
    """创建正向成绩模板（二级表头：环节-方法）"""
    data = _load_relation_json(relation_json_path)
    links = data.get("links", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "正向成绩模板"

    # Row 1: link names, Row 2: method names
    ws.cell(row=1, column=1, value="姓名")
    ws.cell(row=2, column=1, value="姓名")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    method_columns = []
    for link in links:
        methods = link.get("methods", [])
        if not methods:
            methods = [{"name": "无"}]
        start_col = col
        for method in methods:
            ws.cell(row=2, column=col, value=method.get("name", ""))
            method_columns.append(col)
            col += 1
        end_col = col - 1
        ws.cell(row=1, column=start_col, value=link.get("name", ""))
        if end_col >= start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    _apply_header_style(ws, 1, col - 1)
    _apply_header_style(ws, 2, col - 1)

    _append_blank_rows(ws, student_count, col - 1, start_row=3)
    _protect_sheet(ws, editable_cols=list(range(1, col)))

    _apply_grid_borders(ws, ws.max_row, ws.max_column)

    output_path = os.path.join(_ensure_outputs_dir(base_dir), "正向成绩模板.xlsx")
    wb.save(output_path)
    return output_path


def _append_blank_rows(ws, student_count: int, col_count: int, start_row: int = 2):
    for i in range(student_count):
        row_idx = start_row + i
        for col in range(1, col_count + 1):
            ws.cell(row=row_idx, column=col, value="")


def _apply_header_style(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _protect_sheet(ws, editable_cols: List[int], header_rows: int = 2):
    ws.protection.sheet = True
    ws.protection.enable()
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if row <= header_rows:
                cell.protection = Protection(locked=True)
                continue
            cell.protection = Protection(locked=(col not in editable_cols))


def _apply_grid_borders(ws, max_row: int, max_col: int):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = border