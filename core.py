import os
import numpy as np
import pandas as pd
import json
import re
import requests
from typing import List, Dict,Callable, Optional
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from apply_noise import GradeReverseEngine
from utils import normalize_score, get_grade_level, calculate_final_score, calculate_achievement_level, adjust_column_widths
import time

class GradeProcessor:
    def __init__(self, course_name_input, num_objectives_input, weight_inputs, usual_ratio_input,
                 midterm_ratio_input, final_ratio_input, status_label, input_file,
                 course_description="", objective_requirements=None, relation_payload=None):
        self.course_name_input = course_name_input
        self.num_objectives_input = num_objectives_input
        self.weight_inputs = weight_inputs
        self.usual_ratio_input = usual_ratio_input
        self.midterm_ratio_input = midterm_ratio_input
        self.final_ratio_input = final_ratio_input
        self.status_label = status_label
        self.input_file = input_file
        self.course_description = course_description
        self.objective_requirements = objective_requirements or []
        self.previous_achievement_data = None
        self.api_key = None
        self.relation_payload = relation_payload or {}
        self.noise_config = None
        self.reverse_engine = GradeReverseEngine()

    def test_deepseek_api(self, api_key: str) -> str:
        """测试 DeepSeek API 连接"""
        url = "https://api.deepseek.com/v1/chat/completions"
        api_key = api_key.strip().strip('<').strip('>')
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": "测试连接"}
            ],
            "temperature": 0.7,
            "top_p": 1,
            "max_tokens": 10,
            "stream": False
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            response.raise_for_status()
            return "连接成功"
        except requests.RequestException as e:
            error_message = f"连接失败: {str(e)}"
            if hasattr(e, 'response') and e.response is not None:
                error_message += f"\n服务器返回: {e.response.text}"
            return error_message

    def call_deepseek_api(self, prompt: str) -> str:
        """调用 DeepSeek API 获取答案，包含重试与超时处理。"""
        if not self.api_key:
            return "请先设置API Key"

        url = "https://api.deepseek.com/v1/chat/completions"
        api_key = self.api_key.strip().strip("<").strip(">")
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }

        max_tokens = 600
        match = re.search(r"接近(\d+)字", prompt)
        if match:
            try:
                word_limit = int(match.group(1))
                max_tokens = max(200, min(1500, word_limit * 4))
            except ValueError:
                max_tokens = 600

        payload = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": "You are a helpful assistant specializing in course analysis and improvement."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.7,
            "top_p": 1,
            "max_tokens": max_tokens,
            "stream": False,
        }

        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = requests.post(url, headers=headers, json=payload, timeout=30)
                response.raise_for_status()
                return response.json()["choices"][0]["message"]["content"].strip()
            except requests.Timeout:
                if attempt < max_retries - 1:
                    print(f"API 调用超时，正在重试（第 {attempt + 1}/{max_retries} 次）...")
                    time.sleep(2)
                    continue
                return "API 调用超时，请检查网络连接或稍后重试（可能需要使用VPN或代理访问 api.deepseek.com）"
            except requests.RequestException as e:
                error_message = f"API 调用失败: {str(e)}"
                if hasattr(e, "response") and e.response is not None:
                    error_message += f"\n服务端返回: {e.response.text}"
                if attempt < max_retries - 1:
                    print(f"API 调用失败，正在重试（第 {attempt + 1}/{max_retries} 次）...")
                    time.sleep(2)
                    continue
                return error_message
            except (KeyError, IndexError):
                return "API 返回格式错误，无法解析结果"

    def calculate_score_bounds(self, target_score: float, spread_mode: str) -> tuple:
        spread_ranges = {'large': 23, 'medium': 13, 'small': 8}
        base_spread = spread_ranges[spread_mode]

        if target_score < 40:
            spread = min(base_spread, target_score + 5)
        else:
            spread = base_spread

        min_bound = max(0.0, target_score - spread)
        max_bound = min(99.0, target_score + spread)
        
        return min_bound, max_bound

    def generate_initial_scores(self, target, n, min_bound, max_bound, dist_type):
        """生成初始整数分数，分段体现正态分布或偏态分布"""
        scores = np.zeros(n, dtype=int)
        mean = target
        std = (max_bound - min_bound) / 2

        if dist_type == 'normal':
            segments = [
                (mean - 3*std, mean - 2*std, 0.10),
                (mean - 2*std, mean - std, 0.20),
                (mean - std, mean + std, 0.40),
                (mean + std, mean + 2*std, 0.20),
                (mean + 2*std, mean + 3*std, 0.10)
            ]
        elif dist_type == 'left_skewed':
            segments = [
                (min_bound, mean - std, 0.1),
                (mean - std, mean, 0.2),
                (mean, mean + std, 0.4),
                (mean + std, max_bound, 0.3)
            ]
        elif dist_type == 'right_skewed':
            segments = [
                (min_bound, mean - std, 0.3),
                (mean - std, mean, 0.4),
                (mean, mean + std, 0.2),
                (mean + std, max_bound, 0.1)
            ]
        else:
            segments = [(min_bound, max_bound, 1.0)]

        remaining_indices = list(range(n))
        for segment_min, segment_max, proportion in segments:
            num_scores = max(1, int(round(proportion * n)))
            if num_scores == 0:
                continue
            low = max(int(segment_min), int(min_bound))
            high = min(int(segment_max), int(max_bound)) + 1
            if low >= high:
                low = max(int(min_bound), int(segment_min - 1))
                high = min(int(max_bound) + 1, int(segment_max + 1))
                if low >= high:
                    low = int(min_bound)
                    high = int(max_bound) + 1
            chosen_indices = np.random.choice(remaining_indices, min(num_scores, len(remaining_indices)), replace=False)
            for idx in chosen_indices:
                try:
                    scores[idx] = np.random.randint(low, high)
                except ValueError as e:
                    print(f"Error in np.random.randint: low={low}, high={high}, error={str(e)}")
                    scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)
                remaining_indices.remove(idx)

        for idx in remaining_indices:
            scores[idx] = np.random.randint(int(min_bound), int(max_bound) + 1)

        return scores

    def adjust_scores(self, scores, target, weights, min_bound, max_bound, dist_type):
        """逐步调整分数以满足加权和约束，同时保留分布形态"""
        weights_array = np.array(weights)
        scores = np.array(scores, dtype=float)
        max_attempts = 1000
        attempt = 0

        while attempt < max_attempts:
            current_sum = np.sum(scores * weights_array)
            diff = target - current_sum

            if abs(diff) <= 0.1:
                break

            indices = np.argsort(-weights_array)
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    adjustment = diff / weight
                    if adjustment > 0:
                        adjustment = max(1, int(adjustment))
                    else:
                        adjustment = min(-1, int(adjustment))
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

            attempt += 1

        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            indices = np.argsort(weights_array)
            for idx in indices:
                weight = weights[idx]
                if weight > 0:
                    adjustment = round(diff / weight, 1)
                    new_score = scores[idx] + adjustment
                    if min_bound <= new_score <= max_bound:
                        scores[idx] = new_score
                        break

        current_sum = np.sum(scores * weights_array)
        diff = target - current_sum
        if abs(diff) > 0.1:
            print(f"Warning: Final weighted sum deviation {abs(diff):.2f} exceeds 0.1 for target {target}")

        return scores

    def generate_weighted_scores(self, target_sum: float, weights: List[float], all_scores: List[List[float]], 
                                spread_mode: str = 'medium', distribution: str = 'uniform') -> List[float]:
        """
        基于分布模式和跨度范围生成成绩，确保加权和偏差 ≤ 0.1。
        """
        n = len(weights)

        if abs(target_sum) < 0.0001:
            return np.zeros(n).tolist()

        min_bound, max_bound = self.calculate_score_bounds(target_sum, spread_mode)
        scores = self.generate_initial_scores(target_sum, n, min_bound, max_bound, distribution)
        optimized_scores = self.adjust_scores(scores, target_sum, weights, min_bound, max_bound, distribution)

        print(f"Generated scores: {optimized_scores.tolist()}")
        print(f"Distribution - Mean: {np.mean(optimized_scores):.2f}, Std: {np.std(optimized_scores):.2f}")

        return optimized_scores.tolist()

    def process_grades(self, num_objectives, weights, usual_ratio, midterm_ratio, final_ratio, 
                      spread_mode='medium', distribution='uniform',progress_callback: Optional[Callable[[int], None]] = None):
        """处理成绩数据"""
        course_name = self.course_name_input.text()
        if not course_name:
            raise ValueError("请输入课程名称")
            
        output_dir = os.path.dirname(self.input_file)
        detail_output = os.path.join(output_dir, f'{course_name}成绩单详情.xlsx')

        try:
            df = pd.read_excel(self.input_file)
        except Exception as e:
            raise ValueError(f"无法读取输入文件: {str(e)}")

        required_columns = ['学生姓名', '平时成绩', '期中成绩', '期末成绩', '总和']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"输入文件缺少以下必需列：{', '.join(missing_columns)}。请确保文件包含：{', '.join(required_columns)}")
        
        result_data = []
        all_usual_scores = [[] for _ in range(num_objectives)]
        all_midterm_scores = [[] for _ in range(num_objectives)]
        all_final_scores = [[] for _ in range(num_objectives)]
        
        for idx, row in df.iterrows():
            self.status_label.setText(f"正在处理第 {idx+1}/{len(df)} 个学生的成绩...")
            if progress_callback:
                progress_callback(idx)  # 调用进度回调
            name = row['学生姓名']
            total_usual = row['平时成绩']
            total_midterm = row['期中成绩']
            total_final = row['期末成绩']
            total_score = row['总和']
            
            try:
                usual_scores = self.generate_weighted_scores(total_usual, weights, all_usual_scores, spread_mode, distribution)
                midterm_scores = self.generate_weighted_scores(total_midterm, weights, all_midterm_scores, spread_mode, distribution)
                final_scores = self.generate_weighted_scores(total_final, weights, all_final_scores, spread_mode, distribution)
            except Exception as e:
                print(f"Error generating scores for student {name}: {str(e)}")
                raise

            for i in range(num_objectives):
                all_usual_scores[i].append(usual_scores[i])
                all_midterm_scores[i].append(midterm_scores[i])
                all_final_scores[i].append(final_scores[i])
            
            for i in range(num_objectives):
                score = calculate_final_score(
                    usual_scores[i], midterm_scores[i], final_scores[i],
                    usual_ratio, midterm_ratio, final_ratio
                )
                
                result_data.append({
                    '学生姓名': name,
                    '课程目标': i + 1,
                    '平时成绩': usual_scores[i],
                    '期中成绩': midterm_scores[i],
                    '期末成绩': final_scores[i],
                    '权重': weights[i],
                    '平时成绩占比': usual_ratio,
                    '期中成绩占比': midterm_ratio,
                    '期末成绩占比': final_ratio,
                    '分数': score,
                    '等级': get_grade_level(score)
                })
            
            final_total_score = calculate_final_score(
                total_usual, total_midterm, total_final,
                usual_ratio, midterm_ratio, final_ratio
            )
            
            result_data.append({
                '学生姓名': name,
                '课程目标': '总和',
                '平时成绩': total_usual,
                '期中成绩': total_midterm,
                '期末成绩': total_final,
                '权重': sum(weights),
                '平时成绩占比': usual_ratio,
                '期中成绩占比': midterm_ratio,
                '期末成绩占比': final_ratio,
                '分数': final_total_score,
                '等级': get_grade_level(final_total_score)
            })
        
        result_df = pd.DataFrame(result_data)
        
        try:
            with pd.ExcelWriter(detail_output, engine='openpyxl') as writer:
                # 表格从第 1 行开始写入（startrow=0），移除课程简介
                start_row = 0
                result_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=start_row)
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置表格内容居中（包括表头）
                table_start_row = start_row + 1
                table_end_row = table_start_row + len(result_df)
                num_columns = len(result_df.columns)
                for row in range(table_start_row, table_end_row + 1):
                    for col in range(1, num_columns + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 标记“总和”行为黄色
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row_idx, row_data in result_df.iterrows():
                    if row_data['课程目标'] == '总和':
                        excel_row = table_start_row + row_idx + 1  # 转换为 Excel 行号
                        for col in range(1, num_columns + 1):
                            cell = worksheet.cell(row=excel_row, column=col)
                            cell.fill = yellow_fill
                
                # 先调整列宽，再进行合并操作
                adjust_column_widths(worksheet)
                
                # 强制设置姓名列（A 列）宽度为 8 个字符
                worksheet.column_dimensions['A'].width = 8
                
                # 合并姓名列相同的单元格
                current_name = None
                merge_start_row = table_start_row + 1  # 跳过表头行
                for i, name in enumerate(result_df['学生姓名'], start=merge_start_row):
                    if name != current_name:
                        if current_name is not None and merge_start_row != i:
                            worksheet.merge_cells(f'A{merge_start_row}:A{i-1}')
                        current_name = name
                        merge_start_row = i
                
                if merge_start_row != i:
                    worksheet.merge_cells(f'A{merge_start_row}:A{i}')
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        overall_achievement = self.generate_objective_analysis_report(result_df, course_name, weights, usual_ratio, midterm_ratio, final_ratio)
        return overall_achievement

    def generate_objective_analysis_report(self, result_df: pd.DataFrame, course_name: str, weights, usual_ratio, midterm_ratio, final_ratio) -> float:
        """生成课程目标达成度分析报告"""
        output_dir = os.path.dirname(self.input_file)
        analysis_output = os.path.join(output_dir, f'{course_name}课程目标达成度分析表.xlsx')
        
        objectives = sorted([i for i in result_df['课程目标'].unique() if isinstance(i, int)])
        
        analysis_data = []
        
        exam_types = [
            ('平时考核\n(A)', '平时成绩'),
            ('期中考核\n(B)', '期中成绩'),
            ('期末考核\n(C)', '期末成绩')
        ]
        
        weights_dict = {f'课程目标{obj}': round(w * 100, 3) for obj, w in zip(objectives, weights)}
        
        m_values = {}
        
        for exam_name, score_column in exam_types:
            avg_scores = {}
            score_ratios = {}
            
            for obj in objectives:
                obj_scores = result_df[
                    (result_df['课程目标'] == obj)
                ][score_column].tolist()
                
                if obj_scores:
                    avg_scores[f'课程目标{obj}'] = round(np.mean(obj_scores), 1)
                    score_ratios[f'课程目标{obj}'] = 100
            
            analysis_data.extend([
                {
                    '考核环节': exam_name,
                    '指标类型': '平均分',
                    **avg_scores
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分值/满分\n(S)',
                    **score_ratios
                },
                {
                    '考核环节': exam_name,
                    '指标类型': '分权重 (K)',
                    **weights_dict
                }
            ])
        
        m_row = {'考核环节': '课程分目标达成度\n(M)', '指标类型': ''}
        for obj in objectives:
            usual_avg = analysis_data[0].get(f'课程目标{obj}', 0)
            midterm_avg = analysis_data[3].get(f'课程目标{obj}', 0)
            final_avg = analysis_data[6].get(f'课程目标{obj}', 0)
            m = usual_avg * usual_ratio + midterm_avg * midterm_ratio + final_avg * final_ratio
            m_row[f'课程目标{obj}'] = round(m, 1)
            m_values[obj] = m
        
        analysis_data.append(m_row)
        
        z_row = {'考核环节': '课程分目标总权重\n(Z)', '指标类型': ''}
        for obj in objectives:
            z_row[f'课程目标{obj}'] = weights_dict[f'课程目标{obj}']
        analysis_data.append(z_row)
        
        total_achievement = sum(m_values[obj] * weights[obj-1] for obj in objectives)
        total_achievement = round(total_achievement, 1)
        total_row = {'考核环节': '课程总目标达成度', '指标类型': ''}
        for obj in objectives:
            total_row[f'课程目标{obj}'] = total_achievement if obj == objectives[0] else ''
        analysis_data.append(total_row)
        
        columns = ['考核环节', '指标类型'] + [f'课程目标{i}' for i in objectives]
        analysis_df = pd.DataFrame(analysis_data, columns=columns)
        
        try:
            with pd.ExcelWriter(analysis_output, engine='openpyxl') as writer:
                analysis_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置“课程目标”列（从 C 列开始）内容居中
                num_rows = len(analysis_df) + 1  # 包括表头
                for col in range(3, 3 + len(objectives)):  # C 列到 C+len(objectives)-1 列
                    for row in range(1, num_rows + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                current_exam = None
                start_row = 2
                for i, exam in enumerate(analysis_df['考核环节'], start=2):
                    if exam != current_exam:
                        if current_exam is not None and start_row != i:
                            worksheet.merge_cells(f'A{start_row}:A{i-1}')
                        current_exam = exam
                        start_row = i
                
                if start_row != i:
                    worksheet.merge_cells(f'A{start_row}:A{i}')
                
                m_row_idx = len(analysis_df) - 2 + 1
                worksheet.merge_cells(f'A{m_row_idx}:B{m_row_idx}')
                
                z_row_idx = len(analysis_df) - 1 + 1
                worksheet.merge_cells(f'A{z_row_idx}:B{z_row_idx}')
                
                total_row_idx = len(analysis_df) + 1
                worksheet.merge_cells(f'A{total_row_idx}:B{total_row_idx}')
                worksheet.merge_cells(f'C{total_row_idx}:{chr(ord("C") + len(objectives) - 1)}{total_row_idx}')
                worksheet[f'C{total_row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
                
                adjust_column_widths(worksheet)
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

        return total_achievement


    def set_noise_config(self, config: dict):
        """\u8bbe\u7f6e\u566a\u58f0\u914d\u7f6e"""
        self.noise_config = config or None

    def set_relation_payload(self, payload: dict):
        """\u8bbe\u7f6e\u8bfe\u7a0b\u8003\u6838\u4e0e\u76ee\u6807\u5bf9\u5e94\u5173\u7cfb"""
        self.relation_payload = payload or {}


    def _safe_filename(self, name: str) -> str:
        """Sanitize filename for Windows paths."""
        if not name:
            return "\u672a\u547d\u540d"
        safe = re.sub(r'[\\/:*"<>|?\\r\\n\\t]', "_", str(name)).strip()
        return safe or "\u672a\u547d\u540d"

    def _get_links(self):
        payload = self.relation_payload or {}
        return payload.get("links", [])

    def _normalize_weights(self, weights):
        total = sum(weights)
        if total <= 0:
            return [0 for _ in weights]
        return [w / total for w in weights]

    def _validate_forward_headers(self, file_path: str):
        """\u6821\u9a8c\u6b63\u5411\u6a21\u677f\u8868\u5934\u662f\u5426\u4e0e\u5173\u7cfb\u8868\u4e00\u81f4"""
        if not self.relation_payload:
            raise ValueError("\u8bf7\u5148\u586b\u5199\u8bfe\u7a0b\u8003\u6838\u4e0e\u8bfe\u7a0b\u76ee\u6807\u5bf9\u5e94\u5173\u7cfb")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        max_col = ws.max_column

        def header_value(row, col):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                return str(cell.value).strip()
            for rng in ws.merged_cells.ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    v = ws.cell(rng.min_row, rng.min_col).value
                    return str(v).strip() if v is not None else ""
            return ""

        row1 = [header_value(1, c) for c in range(1, max_col + 1)]
        row2 = [header_value(2, c) for c in range(1, max_col + 1)]

        if not row2:
            raise ValueError("\u6b63\u5411\u6a21\u677f\u7b2c1\u5217\u5fc5\u987b\u4e3a\u201c\u59d3\u540d\u201d")
        if row2[0] != "\u59d3\u540d":
            if not (row2[0] == "" and row1[0] == "\u59d3\u540d"):
                raise ValueError("\u6b63\u5411\u6a21\u677f\u7b2c1\u5217\u5fc5\u987b\u4e3a\u201c\u59d3\u540d\u201d")

        links = self._get_links()
        expected_methods = []
        expected_links = []
        for link in links:
            methods = link.get("methods", [])
            if not methods:
                methods = [{"name": "\u65e0"}]
            for m in methods:
                expected_methods.append((m.get("name") or "\u65e0").strip())
                expected_links.append((link.get("name") or "").strip())

        actual_methods = [str(v).strip() for v in row2[1:1+len(expected_methods)]]
        if actual_methods != expected_methods:
            raise ValueError("\u6b63\u5411\u6a21\u677f\u4e8c\u7ea7\u8868\u5934\u4e0e\u5173\u7cfb\u8868\u4e0d\u4e00\u81f4\uff0c\u8bf7\u91cd\u65b0\u4e0b\u8f7d\u6a21\u677f")

        actual_links = [str(v).strip() for v in row1[1:1+len(expected_links)]]
        if actual_links != expected_links:
            raise ValueError("\u6b63\u5411\u6a21\u677f\u4e00\u7ea7\u8868\u5934\u4e0e\u5173\u7cfb\u8868\u4e0d\u4e00\u81f4\uff0c\u8bf7\u91cd\u65b0\u4e0b\u8f7d\u6a21\u677f")

    def _validate_reverse_headers(self, df: pd.DataFrame):
        """\u6821\u9a8c\u9006\u5411\u6a21\u677f\u8868\u5934"""
        links = self._get_links()
        if links:
            expected = ["\u59d3\u540d"] + [link.get("name", "").strip() for link in links]
        else:
            expected = ["\u59d3\u540d", "\u5e73\u65f6\u8003\u6838", "\u671f\u4e2d\u8003\u6838", "\u671f\u672b\u8003\u6838"]

        missing = [c for c in expected if c not in df.columns]
        if missing:
            raise ValueError(f"\u9006\u5411\u6a21\u677f\u7f3a\u5c11\u8868\u5934: {', '.join(missing)}")

    def process_forward_grades(self, spread_mode='medium', distribution='uniform'):
        """\u6b63\u5411\u6210\u7ee9\u5bfc\u5165\u4e0e\u6821\u9a8c\uff0c\u8f93\u51fa\u8be6\u60c5\u6210\u7ee9\u660e\u7ec6\u8868"""
        self._validate_forward_headers(self.input_file)
        df = pd.read_excel(self.input_file, header=1)
        df = df.fillna(0)

        # ??????????????????
        cols = list(df.columns)
        if cols:
            first = str(cols[0]) if cols[0] is not None else ""
            if first.startswith("Unnamed") or first.strip() == "" or first == "nan":
                cols[0] = "\u59d3\u540d"
                df.columns = cols

        if "\u59d3\u540d" not in df.columns:
            raise ValueError("\u6b63\u5411\u6a21\u677f\u7f3a\u5c11\u201c\u59d3\u540d\u201d\u5217")

        links = self._get_links()
        obj_count = int(self.relation_payload.get("objectives_count", 0) or 0)
        if obj_count <= 0:
            obj_keys = set()
            for link in links:
                for method in link.get("methods", []):
                    obj_keys.update(method.get("supports", {}).keys())
            obj_count = len(obj_keys)
        obj_keys = [f"\u8bfe\u7a0b\u76ee\u6807{i+1}" for i in range(obj_count)]
        obj_headers = [f"\u76ee\u6807{i+1}" for i in range(obj_count)]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u6210\u7ee9\u660e\u7ec6"

        header = ["\u59d3\u540d", "\u8003\u6838\u73af\u8282", "\u8003\u6838\u65b9\u5f0f"] + obj_headers + ["\u5c0f\u8ba1", "\u5408\u8ba1", "\u7b49\u7ea7"]
        ws.append(header)

        def format_link_label(name, ratio):
            pct = int(round(ratio * 100))
            if "\u8003\u6838" in name:
                base = name.replace("\u8003\u6838", "")
                return f"{base}\n\u8003\u6838\n({pct}%)"
            return f"{name}\n({pct}%)"

        def grade_label(score):
            if score >= 90:
                return "\u4f18\u79c0"
            if score >= 80:
                return "\u826f\u597d"
            if score >= 70:
                return "\u4e2d\u7b49"
            if score >= 60:
                return "\u53ca\u683c"
            return "\u4e0d\u53ca\u683c"

        row_cursor = 2
        total_scores = []

        for _, row in df.iterrows():
            name = row.get("\u59d3\u540d")
            if pd.isna(name) or str(name).strip() == "":
                continue
            student_start = row_cursor
            total_score = 0.0
            total_obj_scores = [0.0 for _ in range(obj_count)]

            for link in links:
                link_name = link.get("name", "")
                link_ratio = float(link.get("ratio", 0))
                methods = link.get("methods", []) or [{"name": "\u65e0", "supports": {}, "subtotal": 1.0}]

                link_label = format_link_label(link_name, link_ratio)
                link_start = row_cursor
                link_obj_scores = [0.0 for _ in range(obj_count)]
                link_score = 0.0

                for idx, m in enumerate(methods):
                    m_name = m.get("name", "\u65e0")
                    score = row.get(m_name, 0)
                    try:
                        score = float(score)
                    except Exception:
                        score = 0.0

                    supports = m.get("supports", {}) or {}
                    support_vals = [float(supports.get(k, 0)) for k in obj_keys]
                    # ??????????????
                    obj_scores = [score * v for v in support_vals]

                    for i, v in enumerate(obj_scores):
                        link_obj_scores[i] += v

                    method_weight = float(m.get("subtotal", 0))
                    link_score += score * method_weight

                    method_subtotal = sum(obj_scores)
                    row_values = ["", link_label if row_cursor == link_start else "", m_name]
                    row_values += [round(v, 2) for v in obj_scores]
                    row_values += [round(method_subtotal, 2), "", ""]
                    ws.append(row_values)
                    row_cursor += 1

                # ?????
                total_row = ["", link_label if row_cursor == link_start else "", "\u73af\u8282\u5408\u8ba1"]
                total_row += [round(v, 2) for v in link_obj_scores]
                total_row += ["", round(link_score, 2), ""]
                ws.append(total_row)
                row_cursor += 1

                ws.merge_cells(start_row=link_start, start_column=2, end_row=row_cursor - 1, end_column=2)

                total_score += link_score * link_ratio
                for i, v in enumerate(link_obj_scores):
                    total_obj_scores[i] += v * link_ratio

            grade = grade_label(total_score)
            final_row = ["", "100%", "\u8bfe\u7a0b\u603b\u8bc4"]
            final_row += [round(v, 2) for v in total_obj_scores]
            final_row += ["", round(total_score, 2), grade]
            ws.append(final_row)
            row_cursor += 1

            # ??????????
            ws.cell(row=student_start, column=1, value=name)
            grade_col = len(header)
            ws.cell(row=student_start, column=grade_col, value=grade)

            ws.merge_cells(start_row=student_start, start_column=1, end_row=row_cursor - 1, end_column=1)
            ws.merge_cells(start_row=student_start, start_column=grade_col, end_row=row_cursor - 1, end_column=grade_col)
            total_scores.append(total_score)

        # ????? + ??
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.alignment = align
                cell.border = border


        # ???????
        stats_ws = wb.create_sheet(title="\u8bfe\u7a0b\u6210\u7ee9\u7edf\u8ba1")
        total_count = len(total_scores)
        max_score = round(max(total_scores), 2) if total_scores else 0
        min_score = round(min(total_scores), 2) if total_scores else 0
        avg_score = round(float(np.mean(total_scores)) if total_scores else 0.0, 2)

        grade_bins = [
            (90, 100, "\u4f18\u79c0"),
            (80, 89.999, "\u826f\u597d"),
            (70, 79.999, "\u4e2d\u7b49"),
            (60, 69.999, "\u53ca\u683c"),
            (0, 59.999, "\u4e0d\u53ca\u683c"),
        ]
        counts = []
        ratios = []
        for lo, hi, _ in grade_bins:
            c = sum(1 for s in total_scores if lo <= s <= hi)
            counts.append(c)
            ratios.append(round(c / total_count, 4) if total_count else 0)

        def _fmt_ratio(val):
            try:
                val = float(val)
            except Exception:
                val = 0.0
            pct = val * 100 if val <= 1 else val
            if abs(pct - round(pct)) < 0.01:
                return f"{int(round(pct))}%"
            return f"{pct:.2f}%"

        composition_parts = []
        for link in links:
            name = link.get("name", "")
            ratio = link.get("ratio", 0)
            if name:
                composition_parts.append(f"{name}\uff08{_fmt_ratio(ratio)}\uff09")
        composition_text = " + ".join(composition_parts)

        stats_ws.append(["\u6210\u7ee9\u6784\u6210", composition_text, "", "", "", ""])
        stats_ws.merge_cells("B1:F1")
        stats_ws.append(["\u6700\u9ad8\u6210\u7ee9", max_score, "\u6700\u4f4e\u6210\u7ee9", min_score, "\u5e73\u5747\u6210\u7ee9", avg_score])
        stats_ws.append([
            "\u6210\u7ee9\u7b49\u7ea7",
            "90-100\n(\u4f18\u79c0)",
            "80-89\n(\u826f\u597d)",
            "70-79\n(\u4e2d\u7b49)",
            "60-69\n(\u53ca\u683c)",
            "<60\n(\u4e0d\u53ca\u683c)",
        ])
        stats_ws.append(["\u4eba\u6570"] + counts)
        stats_ws.append(["\u5360\u8003\u6838\u4eba\u6570\u7684\u6bd4\u4f8b"] + [f"{r*100:.2f}%" for r in ratios])

        # ??
        stat_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        stat_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in stats_ws.iter_rows(min_row=1, max_row=stats_ws.max_row, min_col=1, max_col=stats_ws.max_column):
            for cell in r:
                cell.alignment = stat_align
                cell.border = stat_border

        stats_ws.row_dimensions[3].height = 36
        stats_ws.column_dimensions["A"].width = 18
        for col in ["B", "C", "D", "E", "F"]:
            stats_ws.column_dimensions[col].width = 14
        # EVAL_TABLE_5
        eval_wb = openpyxl.Workbook()
        eval_ws = eval_wb.active
        eval_ws.title = "\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u8bc4\u4ef7\u7ed3\u679c"
        eval_headers = [
            "\u8bfe\u7a0b\u5206\u76ee\u6807",
            "\u8003\u6838\u73af\u8282",
            "\u5206\u6743\u91cd",
            "\u5206\u503c/\u6ee1\u5206",
            "\u5b66\u751f\u5b9e\u9645\u5f97\u5206\u5e73\u5747\u5206",
            "\u5206\u76ee\u6807\u8fbe\u6210\u503c",
            "\u4e0a\u4e00\u8f6e\u6559\u5b66\u5206\u76ee\u6807\u8fbe\u6210\u503c",
        ]
        eval_ws.append(eval_headers)

        # \u8ba1\u7b97\u5404\u8003\u6838\u65b9\u5f0f\u5e73\u5747\u5206\uff08\u57fa\u4e8e\u5bfc\u5165\u6210\u7ee9\uff09
        method_avgs = {}
        for link in links:
            for m in link.get("methods", []) or []:
                m_name = m.get("name")
                if not m_name:
                    continue
                if m_name in df.columns:
                    try:
                        method_avgs[m_name] = float(df[m_name].mean())
                    except Exception:
                        method_avgs[m_name] = 0.0
                else:
                    method_avgs[m_name] = 0.0

        prev_data = self.previous_achievement_data or {}
        current_achievement = {}
        total_obj_weight = 0.0
        total_obj_actual = 0.0

        row_cursor = 2
        for idx, obj_key in enumerate(obj_keys):
            obj_name = f"\u8bfe\u7a0b\u76ee\u6807{idx + 1}"
            obj_start = row_cursor
            obj_weight_sum = 0.0
            obj_actual_sum = 0.0

            for link in links:
                link_name = link.get("name", "")
                if "\u5e73\u65f6" in link_name:
                    display_link = "\u5e73\u65f6\u6210\u7ee9"
                elif "\u671f\u4e2d" in link_name:
                    display_link = "\u671f\u4e2d\u8003\u6838"
                elif "\u671f\u672b" in link_name:
                    display_link = "\u671f\u672b\u8003\u6838"
                else:
                    display_link = link_name

                link_ratio = float(link.get("ratio", 0))
                methods = link.get("methods", []) or []

                support_sum = 0.0
                actual_sum = 0.0
                for m in methods:
                    supports = m.get("supports", {}) or {}
                    weight = float(supports.get(obj_key, 0))
                    support_sum += weight
                    m_name = m.get("name")
                    m_avg = float(method_avgs.get(m_name, 0))
                    actual_sum += m_avg * weight

                target_weight = link_ratio * 100.0 * support_sum
                actual_score = link_ratio * actual_sum

                obj_weight_sum += target_weight
                obj_actual_sum += actual_score

                eval_ws.append([
                    obj_name if row_cursor == obj_start else "",
                    display_link,
                    round(target_weight, 2),
                    100,
                    round(actual_score, 2),
                    "",
                    "",
                ])
                row_cursor += 1

            achievement = round(obj_actual_sum / obj_weight_sum, 4) if obj_weight_sum > 0 else 0
            current_achievement[obj_name] = achievement
            prev_val = prev_data.get(obj_name, 0) if prev_data else 0
            prev_val = 0 if prev_val is None else prev_val

            eval_ws.cell(row=obj_start, column=6, value=achievement)
            eval_ws.cell(row=obj_start, column=7, value=prev_val)

            if row_cursor - 1 > obj_start:
                eval_ws.merge_cells(start_row=obj_start, start_column=1, end_row=row_cursor - 1, end_column=1)
                eval_ws.merge_cells(start_row=obj_start, start_column=6, end_row=row_cursor - 1, end_column=6)
                eval_ws.merge_cells(start_row=obj_start, start_column=7, end_row=row_cursor - 1, end_column=7)

            total_obj_weight += obj_weight_sum
            total_obj_actual += obj_actual_sum

        total_attainment = round(total_obj_actual / total_obj_weight, 4) if total_obj_weight > 0 else 0
        current_achievement["总达成度"] = total_attainment
        self.current_achievement = current_achievement
        expected_attainment = 0.7
        prev_total = 0
        for key in ["\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c", "\u8bfe\u7a0b\u603b\u76ee\u6807", "\u8bfe\u7a0b\u603b\u8fbe\u6210\u503c", "total_value"]:
            if key in prev_data:
                prev_total = prev_data.get(key, 0) or 0
                break

        def _append_summary(label, value=None, prev=None):
            display_val = value if value is not None else (prev if prev is not None else 0)
            eval_ws.append([label, "", "", "", "", display_val, ""])
            row_idx = eval_ws.max_row
            eval_ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            eval_ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)

        _append_summary("\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c", total_attainment)
        _append_summary("\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u671f\u671b\u503c", expected_attainment)
        _append_summary("\u4e0a\u4e00\u8f6e\u6559\u5b66\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c", None, prev_total)

        eval_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        eval_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in eval_ws.iter_rows(min_row=1, max_row=eval_ws.max_row, min_col=1, max_col=eval_ws.max_column):
            for cell in r:
                cell.alignment = eval_align
                cell.border = eval_border

        eval_ws.column_dimensions["A"].width = 14
        eval_ws.column_dimensions["B"].width = 12
        eval_ws.column_dimensions["C"].width = 10
        eval_ws.column_dimensions["D"].width = 12
        eval_ws.column_dimensions["E"].width = 18
        eval_ws.column_dimensions["F"].width = 12
        eval_ws.column_dimensions["G"].width = 16

        output_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), "outputs")
        os.makedirs(output_dir, exist_ok=True)
        safe_name = self._safe_filename(self.course_name_input.text())
        detail_output_path = os.path.join(output_dir, f"{safe_name}\u6210\u7ee9\u660e\u7ec6.xlsx")
        eval_output_path = os.path.join(output_dir, f"{safe_name}\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u8bc4\u4ef7\u7ed3\u679c.xlsx")
        eval_wb.save(eval_output_path)
        wb.save(detail_output_path)

        return round(float(np.mean(total_scores)) if total_scores else 0.0, 2)

    def process_reverse_grades(self, spread_mode='medium', distribution='uniform'):
        """\u9006\u5411\u6210\u7ee9\u5bfc\u5165\u4e0e\u751f\u6210\u660e\u7ec6"""
        df = pd.read_excel(self.input_file)
        df = df.fillna(0)
        self._validate_reverse_headers(df)

        links = self._get_links()
        if not links:
            links = [
                {"name": "\u5e73\u65f6\u8003\u6838", "ratio": 0.0, "methods": []},
                {"name": "\u671f\u4e2d\u8003\u6838", "ratio": 0.0, "methods": []},
                {"name": "\u671f\u672b\u8003\u6838", "ratio": 0.0, "methods": []},
            ]

        detail_rows = []
        total_scores = []

        dist_map = {
            "normal": "normal",
            "left_skewed": "left_skewed",
            "right_skewed": "right_skewed",
            "bimodal": "bimodal",
            "discrete": "discrete",
            "uniform": "normal",
        }
        dist_type = dist_map.get(distribution, "normal")

        for _, row in df.iterrows():
            name = row.get("\u59d3\u540d")
            if pd.isna(name) or str(name).strip() == "":
                continue
            row_dict = {"\u59d3\u540d": name}
            total_score = 0.0

            for link in links:
                link_name = link.get("name", "")
                link_ratio = float(link.get("ratio", 0))
                link_score = row.get(link_name, 0)
                try:
                    link_score = float(link_score)
                except Exception:
                    link_score = 0.0

                methods = link.get("methods", []) or [{"name": "\u65e0", "subtotal": 1.0}]
                weights = [float(m.get("subtotal", 0)) for m in methods]
                weights = self._normalize_weights(weights)

                structure = {}
                for m, w in zip(methods, weights):
                    structure[m.get("name", "\u65e0")] = {"weight": w, "type": dist_type}

                if structure and sum(weights) > 0:
                    breakdown = self.reverse_engine.generate_breakdown(
                        link_score,
                        structure,
                        noise_config=self.noise_config,
                    )
                else:
                    breakdown = {m.get("name", "\u65e0"): 0 for m in methods}

                for m in methods:
                    m_name = m.get("name", "\u65e0")
                    row_dict[f"{link_name}-{m_name}"] = breakdown.get(m_name, 0)

                row_dict[link_name] = round(link_score, 2)
                total_score += link_score * link_ratio

            row_dict["\u603b\u8bc4"] = round(total_score, 2)
            total_scores.append(total_score)
            detail_rows.append(row_dict)

        
        

        
        detail_df = pd.DataFrame(detail_rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u6210\u7ee9\u660e\u7ec6"
        for r in dataframe_to_rows(detail_df, index=False, header=True):
            ws.append(r)

        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.alignment = align
                cell.border = border

        stats_ws = wb.create_sheet(title="\u8bfe\u7a0b\u6210\u7ee9\u7edf\u8ba1")
        total_count = len(total_scores)
        max_score = round(max(total_scores), 2) if total_scores else 0
        min_score = round(min(total_scores), 2) if total_scores else 0
        avg_score = round(float(np.mean(total_scores)) if total_scores else 0.0, 2)

        grade_bins = [
            (90, 100, "\u4f18\u79c0"),
            (80, 89.999, "\u826f\u597d"),
            (70, 79.999, "\u4e2d\u7b49"),
            (60, 69.999, "\u53ca\u683c"),
            (0, 59.999, "\u4e0d\u53ca\u683c"),
        ]
        counts = []
        ratios = []
        for lo, hi, _ in grade_bins:
            c = sum(1 for s in total_scores if lo <= s <= hi)
            counts.append(c)
            ratios.append(round(c / total_count, 4) if total_count else 0)

        def _fmt_ratio(val):
            try:
                val = float(val)
            except Exception:
                val = 0.0
            pct = val * 100 if val <= 1 else val
            if abs(pct - round(pct)) < 0.01:
                return f"{int(round(pct))}%"
            return f"{pct:.2f}%"

        composition_parts = []
        for link in links:
            name = link.get("name", "")
            ratio = link.get("ratio", 0)
            if name:
                composition_parts.append(f"{name}\uff08{_fmt_ratio(ratio)}\uff09")
        composition_text = " + ".join(composition_parts)

        stats_ws.append(["\u6210\u7ee9\u6784\u6210", composition_text, "", "", "", ""])
        stats_ws.merge_cells("B1:F1")
        stats_ws.append(["\u6700\u9ad8\u6210\u7ee9", max_score, "\u6700\u4f4e\u6210\u7ee9", min_score, "\u5e73\u5747\u6210\u7ee9", avg_score])
        stats_ws.append([
            "\u6210\u7ee9\u7b49\u7ea7",
            "90-100\n(\u4f18\u79c0)",
            "80-89\n(\u826f\u597d)",
            "70-79\n(\u4e2d\u7b49)",
            "60-69\n(\u53ca\u683c)",
            "<60\n(\u4e0d\u53ca\u683c)",
        ])
        stats_ws.append(["\u4eba\u6570"] + counts)
        stats_ws.append(["\u5360\u8003\u6838\u4eba\u6570\u7684\u6bd4\u4f8b"] + [f"{r*100:.2f}%" for r in ratios])

        stat_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        stat_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in stats_ws.iter_rows(min_row=1, max_row=stats_ws.max_row, min_col=1, max_col=stats_ws.max_column):
            for cell in r:
                cell.alignment = stat_align
                cell.border = stat_border

        stats_ws.row_dimensions[3].height = 36
        stats_ws.column_dimensions["A"].width = 18
        for col in ["B", "C", "D", "E", "F"]:
            stats_ws.column_dimensions[col].width = 14


        output_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), "outputs")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{self._safe_filename(self.course_name_input.text())}\u6210\u7ee9\u660e\u7ec6.xlsx")
        wb.save(output_path)

        return round(float(np.mean(total_scores)) if total_scores else 0.0, 2)




    def load_previous_achievement(self, file_path: str) -> None:
        """\u52a0\u8f7d\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6\u8868\uff0c\u63d0\u53d6\u5404\u8bfe\u7a0b\u76ee\u6807\u7684\u5206\u76ee\u6807\u8fbe\u6210\u503c\u4ee5\u53ca\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c\u3002"""
        def _objective_count():
            payload = self.relation_payload or {}
            objectives = payload.get("objectives") if isinstance(payload, dict) else None
            if objectives:
                return len(objectives)
            if self.objective_requirements:
                return len(self.objective_requirements)
            return 5

        def _init_defaults(n):
            data = {f'\u8bfe\u7a0b\u76ee\u6807{i}': 0 for i in range(1, n + 1)}
            data['\u8bfe\u7a0b\u603b\u76ee\u6807'] = 0
            return data

        obj_count = _objective_count()
        if not file_path:
            self.previous_achievement_data = _init_defaults(obj_count)
            return

        try:
            if not os.path.exists(file_path):
                self.previous_achievement_data = _init_defaults(obj_count)
                if self.status_label:
                    self.status_label.setText("\u672a\u627e\u5230\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6\u8868\uff0c\u5df2\u4f7f\u7528\u9ed8\u8ba4\u503c")
                return

            xls = pd.ExcelFile(file_path)
            df = None
            for sheet in xls.sheet_names:
                tmp = pd.read_excel(file_path, sheet_name=sheet)
                cols = [str(c).strip() for c in tmp.columns]
                if "\u8bfe\u7a0b\u5206\u76ee\u6807" in cols or "\u8bfe\u7a0b\u76ee\u6807" in cols:
                    df = tmp
                    break
            if df is None:
                df = pd.read_excel(file_path)

            cols = [str(c).strip() for c in df.columns]
            data = _init_defaults(obj_count)

            if "\u8bfe\u7a0b\u5206\u76ee\u6807" in cols and "\u5206\u76ee\u6807\u8fbe\u6210\u503c" in cols:
                for i in range(1, obj_count + 1):
                    key = f'\u8bfe\u7a0b\u76ee\u6807{i}'
                    rows = df[df["\u8bfe\u7a0b\u5206\u76ee\u6807"].astype(str).str.strip() == key]
                    if not rows.empty:
                        val = rows["\u5206\u76ee\u6807\u8fbe\u6210\u503c"].dropna().tolist()
                        if val:
                            data[key] = float(val[0])
                total_rows = df[df["\u8bfe\u7a0b\u5206\u76ee\u6807"].astype(str).str.strip().isin([
                    "\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c",
                    "\u8bfe\u7a0b\u603b\u76ee\u6807\u8fbe\u6210\u503c",
                    "\u8bfe\u7a0b\u603b\u8fbe\u6210\u503c",
                ])]
                if not total_rows.empty:
                    val = total_rows["\u5206\u76ee\u6807\u8fbe\u6210\u503c"].dropna().tolist()
                    if val:
                        data["\u8bfe\u7a0b\u603b\u76ee\u6807"] = float(val[0])
            elif "\u8bfe\u7a0b\u76ee\u6807" in cols:
                value_col = None
                for cand in [
                    "\u4e0a\u4e00\u5e74\u5ea6\u8fbe\u6210\u5ea6",
                    "\u4e0a\u4e00\u8f6e\u6559\u5b66\u5206\u76ee\u6807\u8fbe\u6210\u503c",
                    "\u5206\u76ee\u6807\u8fbe\u6210\u503c",
                ]:
                    if cand in cols:
                        value_col = cand
                        break
                if value_col:
                    for _, row in df.iterrows():
                        target = str(row["\u8bfe\u7a0b\u76ee\u6807"]).strip()
                        if target in data and isinstance(row[value_col], (int, float)):
                            data[target] = float(row[value_col])
                    total_rows = df[df["\u8bfe\u7a0b\u76ee\u6807"].astype(str).str.strip().isin([
                        "\u8bfe\u7a0b\u76ee\u6807\u8fbe\u6210\u503c",
                        "\u8bfe\u7a0b\u603b\u76ee\u6807\u8fbe\u6210\u503c",
                        "\u8bfe\u7a0b\u603b\u8fbe\u6210\u503c",
                    ])]
                    if not total_rows.empty and isinstance(total_rows[value_col].iloc[0], (int, float)):
                        data["\u8bfe\u7a0b\u603b\u76ee\u6807"] = float(total_rows[value_col].iloc[0])
            self.previous_achievement_data = data
            if self.status_label:
                self.status_label.setText(f"\u5df2\u52a0\u8f7d\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6\u8868: {os.path.basename(file_path)}")
        except Exception as e:
            if self.status_label:
                self.status_label.setText("\u52a0\u8f7d\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6\u8868\u5931\u8d25")
            raise ValueError(f"\u52a0\u8f7d\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6\u8868\u5931\u8d25: {str(e)}")

    def generate_improvement_report(self, current_achievement: Dict[str, float], course_name: str, num_objectives: int, answers=None) -> None:
        """Generate single-column improvement report."""
        output_dir = os.path.join(os.path.abspath(os.path.dirname(__file__)), "outputs")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(
            output_dir,
            f"{self._safe_filename(course_name)}\u8bfe\u7a0b\u5206\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u5206\u6790\u3001\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd.xlsx",
        )

        num_objectives = max(1, int(num_objectives))
        questions = []
        for i in range(1, num_objectives + 1):
            questions.append(f"\u8bfe\u7a0b\u76ee\u6807{i}\uff081\uff09\u8fbe\u6210\u60c5\u51b5\u5206\u6790")
            questions.append(f"\u8bfe\u7a0b\u76ee\u6807{i}\uff082\uff09\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd")

        if answers is None:
            prev_data = self.previous_achievement_data or {}
            current_data = current_achievement or {}
            prev_total = prev_data.get("\u8bfe\u7a0b\u603b\u76ee\u6807", 0)
            current_total = current_data.get("\u603b\u8fbe\u6210\u5ea6", 0)

            context = f"\u8bfe\u7a0b\u7b80\u4ecb: {self.course_description}\n"
            for i, req in enumerate(self.objective_requirements, 1):
                context += f"\u8bfe\u7a0b\u76ee\u6807{i}\u8981\u6c42: {req}\n"
            for i in range(1, num_objectives + 1):
                prev_score = prev_data.get(f"\u8bfe\u7a0b\u76ee\u6807{i}", 0)
                current_score = current_data.get(f"\u8bfe\u7a0b\u76ee\u6807{i}", 0)
                context += f"\u8bfe\u7a0b\u76ee\u6807{i}\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6: {prev_score}\n"
                context += f"\u8bfe\u7a0b\u76ee\u6807{i}\u672c\u5b66\u5e74\u8fbe\u6210\u5ea6: {current_score}\n"
            context += f"\u8bfe\u7a0b\u603b\u76ee\u6807\u4e0a\u4e00\u5b66\u5e74\u8fbe\u6210\u5ea6: {prev_total}\n"
            context += f"\u8bfe\u7a0b\u603b\u76ee\u6807\u672c\u5b66\u5e74\u8fbe\u6210\u5ea6: {current_total}\n"

            cache_file = os.path.join(output_dir, "api_cache.json")
            cached_answers = {}
            if os.path.exists(cache_file):
                try:
                    with open(cache_file, "r", encoding="utf-8") as f:
                        cached_answers = json.load(f)
                except Exception:
                    cached_answers = {}

            answers = []
            total_questions = len(questions)
            for i, question in enumerate(questions):
                if self.status_label:
                    self.status_label.setText(f"\u6b63\u5728\u5904\u7406\u7b2c {i + 1}/{total_questions} \u4e2a\u95ee\u9898...")
                prompt = f"{context}\n\u95ee\u9898: {question}"
                cache_key = f"{course_name}_{question}"
                if cache_key in cached_answers:
                    answers.append(cached_answers[cache_key])
                else:
                    answer = self.call_deepseek_api(prompt)
                    cached_answers[cache_key] = answer
                    answers.append(answer)
                    try:
                        with open(cache_file, "w", encoding="utf-8") as f:
                            json.dump(cached_answers, f, indent=4, ensure_ascii=False)
                    except Exception:
                        pass

        rows = []
        overall_answer = ""
        answer_idx = 0
        if answers and len(answers) == (num_objectives * 2 + 1):
            overall_answer = answers[0]
            answer_idx = 1
        rows.append("\uff08\u4e00\uff09\u603b\u4f53\u60c5\u51b5")
        rows.append(overall_answer or "")
        rows.append("\uff08\u4e8c\uff09\u8bfe\u7a0b\u5206\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u5206\u6790\u3001\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd")
        for i in range(1, num_objectives + 1):
            rows.append(f"{i}.\u8bfe\u7a0b\u76ee\u6807{i}")
            rows.append("\uff081\uff09\u8fbe\u6210\u60c5\u51b5\u5206\u6790\uff1a")
            rows.append(answers[answer_idx] if answers else "")
            answer_idx += 1
            rows.append("\uff082\uff09\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd\uff1a")
            rows.append(answers[answer_idx] if answers else "")
            answer_idx += 1

        heading_texts = {
            "\uff08\u4e00\uff09\u603b\u4f53\u60c5\u51b5",
            "\uff08\u4e8c\uff09\u8bfe\u7a0b\u5206\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u5206\u6790\u3001\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd",
        }
        heading_font = Font(name="\u4eff\u5b8b", size=16, bold=True)
        body_font = Font(name="\u4eff\u5b8b", size=16, bold=False)
        no_border = Border()

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "\u8bfe\u7a0b\u5206\u76ee\u6807\u8fbe\u6210\u60c5\u51b5\u5206\u6790\u3001\u5b58\u5728\u95ee\u9898\u53ca\u6539\u8fdb\u63aa\u65bd"
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 67
            ws.sheet_view.showGridLines = False

            for row_idx, text in enumerate(rows, start=1):
                cell = ws.cell(row=row_idx, column=2)
                if text in heading_texts:
                    cell.value = text
                    cell.font = heading_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    display_text = text if text else ""
                    cell.value = display_text
                    cell.font = body_font
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.border = no_border
            wb.save(output_file)
        except Exception as e:
            print(f"Error writing to Excel: {str(e)}")
            raise

    def store_api_key(self, api_key: str) -> None:
        """\u5b58\u50a8 API Key\u3002"""
        self.api_key = api_key
        if self.status_label:
            self.status_label.setText("\u5df2\u5b58\u50a8API Key")

    def generate_ai_report(self, num_objectives: int, current_achievement: Dict[str, float]) -> None:
        """\u751f\u6210AI\u5206\u6790\u62a5\u544a\u3002"""
        if not self.api_key:
            raise ValueError("\u8bf7\u5148\u8bbe\u7f6eAPI Key")

        course_name = self.course_name_input.text()
        if not course_name:
            if self.status_label:
                self.status_label.setText("\u8bf7\u5148\u8f93\u5165\u8bfe\u7a0b\u540d\u79f0")
            return

        try:
            self.generate_improvement_report(current_achievement, course_name, num_objectives)
            if self.status_label:
                self.status_label.setText("AI\u5206\u6790\u62a5\u544a\u5df2\u751f\u6210")
        except Exception as e:
            if self.status_label:
                self.status_label.setText("\u751f\u6210AI\u5206\u6790\u62a5\u544a\u5931\u8d25")
            raise ValueError(f"\u751f\u6210AI\u5206\u6790\u62a5\u544a\u5931\u8d25: {str(e)}")